import openpyxl
import re

def get_seller_data_from_excel(file_path, sheet_name="Sheet1"):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    seller_list = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            full_name = " ".join(str(cell_value).split())
            
            no_finance_name = re.sub(r'\bFINANCE\b\s*', '', full_name, flags=re.IGNORECASE)
            no_finance_name = " ".join(no_finance_name.split())
            
            matches = re.findall(r'\((.*?)\)', full_name)
            objek_lelang = matches[-1].strip().title() if matches else "Bike"
            
            seller_list.append({
                "no_sub_seller": no_finance_name,
                "nama_sub_seller": full_name,
                "objek_lelang": objek_lelang,
                "kategori_seller": "Gold Seller"
            })
            
    return seller_list