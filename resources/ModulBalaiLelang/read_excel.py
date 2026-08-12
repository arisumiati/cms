import openpyxl


def get_data_edit_lelang_from_excel(file_path, sheet_name="edit lane dan lot"):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = (
        workbook[sheet_name]
        if sheet_name in workbook.sheetnames
        else workbook.active
    )

    data_list = []
    for r in range(2, sheet.max_row + 1):
        alamat = sheet.cell(r, 4).value
        if alamat:
            data_list.append({
                "alamat": " ".join(str(alamat).split()),
                "lane_car": str(sheet.cell(r, 5).value or "").strip().upper(),
                "lot_car": str(sheet.cell(r, 6).value or "").strip(),
                "lane_bike": str(sheet.cell(r, 7).value or "").strip().upper(),
                "lot_bike": str(sheet.cell(r, 8).value or "").strip(),
            })

    return data_list