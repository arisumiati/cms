import openpyxl

def get_alamat_balai_lelang_from_excel(file_path, sheet_name="sheet2"):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    alamat_list = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=5).value
        if cell_value:
            clean_address = " ".join(str(cell_value).split())
            alamat_list.append(clean_address)
            
    return alamat_list

import openpyxl

def get_seller_format_lengkap_from_excel(file_path, sheet_name="Sheet1"):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    seller_list = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value:
            clean_seller = " ".join(str(cell_value).split())
            seller_list.append(clean_seller)
            
    return seller_list